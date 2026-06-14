"""Legger til den sene deltakeren Duc i Drammen sin participants.ts.
Gruppespill-tips leses direkte fra hans separate skjema (data/VM 2026 skjema-Duc-Drammen.xlsx);
krydder-svarene er tolket for hånd (han rotet litt) – se DUC_BONUS under.
Kjør på nytt etter en full regenerering av participants.ts.  py tools/add_duc.py
"""
import openpyxl, re, json
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "data" / "VM 2026 skjema-Duc-Drammen.xlsx"
PT = ROOT / "apps" / "drammen" / "src" / "data" / "participants.ts"
SEP = re.compile(r"\s[–—-]\s")

# Tolkede krydder-svar (q10/q13/q15/q17 står blanke – ikke besvart/ikke i hans skjema).
DUC_BONUS = [
    ("q1", "Norge"),               # svaret lå én rad for høyt (O2) – tolket som VM-vinner
    ("q2", "Messi"),
    ("q3", "Haaland"),
    ("q4", "Jamal"),               # rå tekst (Musiala/Yamal? – fasit-setter avgjør)
    ("q5", "189"),                 # antall mål: brukte 189 (O20); «-5» (O10) var margin-rot
    ("q6", "07:30"),               # skrev «5-10 minutter» → midtpunkt 7:30
    ("q7", ["DR Kongo", "Ghana"]), # «Kongo» → DR Kongo
    ("q8", ["Curaçao", "Tunisia"]),# «Curacao» → Curaçao
    ("q9", "I og L"),              # oppga to grupper
    ("q11", "Norsk Dommer"),
    ("q12", "Samoa"),              # ikke i VM – hans svar beholdt
    ("q14", "Marokko"),
    ("q16", "Ja"),
]

def parse_groups():
    ws = openpyxl.load_workbook(XLSX, data_only=True)["Gruppespill og krydder"]
    group, tips = None, []
    for r in range(1, ws.max_row + 1):
        i = ws.cell(r, 9).value
        l = ws.cell(r, 12).value
        if i is None: continue
        s = str(i).strip()
        mg = re.match(r"Gruppe ([A-L])$", s)
        if mg: group = "GROUP_" + mg.group(1); continue
        if group is None or l is None or str(l).strip() == "": continue
        parts = SEP.split(s)
        if len(parts) != 2: continue
        mr = re.match(r"(\d+)\s*[-–]\s*(\d+)$", str(l).strip())
        if not mr: continue
        tips.append((parts[0].strip(), parts[1].strip(), group, int(mr.group(1)), int(mr.group(2))))
    return tips

def j(v): return json.dumps(v, ensure_ascii=False)

def main():
    tips = parse_groups()
    lines = ["  {", '    name: "Duc",', "    groupTips: ["]
    for h, a, g, hg, ag in tips:
        lines.append(f"      {{ homeTeam: {j(h)}, awayTeam: {j(a)}, group: {j(g)}, homeGoals: {hg}, awayGoals: {ag} }},")
    lines.append("    ],")
    lines.append("    bonusTips: [")
    for qid, ans in DUC_BONUS:
        lines.append(f"      {{ questionId: {j(qid)}, answer: {j(ans)} }},")
    lines += ["    ],", "    knockoutTips: [],", "  },"]
    block = "\n".join(lines)

    content = PT.read_text(encoding="utf-8")
    if '"Duc"' in content:
        print("Duc finnes allerede – avbryter (kjør evt. full regenerering først)."); return
    idx = content.rstrip().rfind("];")
    new = content.rstrip()[:idx] + block + "\n];\n"
    PT.write_text(new, encoding="utf-8")
    print(f"La til Duc med {len(tips)} gruppespill-tips og {len(DUC_BONUS)} krydder-svar.")

if __name__ == "__main__":
    main()
