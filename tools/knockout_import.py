# -*- coding: utf-8 -*-
"""Konverterer et runde-regneark (kamptips + krydder-svar) til Importer-JSON.

Bruk (fra repo-rot):
  py tools/knockout_import.py "data/Kvartfinaler - Alles tips.xlsx" --app alles \
      --round QF --k k4 k5 k6 -o data/qf_import_alles.json

Regnearkformat (samme som «Åttedelsfinaler»-arkene):
  Ark «Resultater»: rad 1 har deltakernavn fra kolonne 4 og utover; deretter én rad
    per kamp med hjemmelag (kol 1), «-» (kol 2), bortelag (kol 3) og «h-b»-tips per
    deltaker. Tomme celler = deltakeren har ikke tippet kampen (hoppes over).
  Ark «Krydder»: rad 1 er header (spørsmål, poeng, deltakernavn fra kolonne 3);
    deretter én rad per spørsmål. Radrekkefølgen mappes til id-ene gitt med --k.

Output-JSON limes inn i admin → Importer (flettes inn i KV, publiseres til alle).
Kamp-orientering sjekkes mot API-oppsettet: står kampen speilvendt i arket,
snus tipsene automatisk (rapporteres).
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent

# apiId per kamp, med lagnavn på norsk slik de skrives i arkene. Runder som IKKE ligger
# her (semifinaler, bronse, finale) hentes live fra API-et når lagene er klare – se
# fetch_fixtures(). R16/QF står innbakt så regresjonstestene virker offline.
FIXTURES: dict[str, dict[int, tuple[str, str]]] = {
    'R16': {
        537375: ('Paraguay', 'Frankrike'),
        537376: ('Canada', 'Marokko'),
        537377: ('Brasil', 'Norge'),
        537378: ('Mexico', 'England'),
        537379: ('Portugal', 'Spania'),
        537380: ('USA', 'Belgia'),
        537381: ('Argentina', 'Egypt'),
        537382: ('Sveits', 'Colombia'),
    },
    'QF': {
        537383: ('Frankrike', 'Marokko'),
        537384: ('Spania', 'Belgia'),
        537385: ('Norge', 'England'),
        537386: ('Argentina', 'Sveits'),
    },
}

# Runde → football-data stage-parameter (for live-henting av oppsett).
ROUND_STAGE = {
    'R32': 'LAST_32',
    'R16': 'LAST_16',
    'QF': 'QUARTER_FINALS',
    'SF': 'SEMI_FINALS',
    'BRONSE': 'THIRD_PLACE',
    'FINALE': 'FINAL',
}


def team_name_map() -> dict[str, str]:
    """Engelsk API-navn → norsk, parset fra den genererte teamNames.ts (alle 48 lag)."""
    src = (REPO_ROOT / 'apps' / 'drammen' / 'src' / 'utils' / 'teamNames.ts').read_text(encoding='utf-8')
    return dict(re.findall(r'"([^"]+)":\s*"([^"]+)"', src))


def fetch_fixtures(round_key: str) -> dict[int, tuple[str, str]]:
    """Hent rundens kampoppsett (apiId + norske lagnavn) live fra football-data.org."""
    import json as _json
    import urllib.request

    env = (REPO_ROOT / 'apps' / 'drammen' / '.env.local').read_text(encoding='utf-8')
    m = re.search(r'FOOTBALL_API_KEY\s*=\s*"?([^"\r\n]+)', env)
    if not m:
        sys.exit('FEIL: fant ikke FOOTBALL_API_KEY i apps/drammen/.env.local (trengs for å hente oppsettet).')
    req = urllib.request.Request(
        f'https://api.football-data.org/v4/competitions/WC/matches?stage={ROUND_STAGE[round_key]}',
        headers={'X-Auth-Token': m.group(1).strip()},
    )
    with urllib.request.urlopen(req) as r:
        matches = _json.load(r).get('matches', [])
    no = team_name_map()
    out: dict[int, tuple[str, str]] = {}
    for match in matches:
        home = match['homeTeam'].get('name')
        away = match['awayTeam'].get('name')
        if not home or not away:
            sys.exit(f'FEIL: {round_key} er ikke trukket ennå (kamp {match["id"]} har TBD-lag). Prøv igjen når runden er klar.')
        out[match['id']] = (no.get(home, home), no.get(away, away))
    if not out:
        sys.exit(f'FEIL: API-et returnerte ingen kamper for {round_key}.')
    return out

# Regneark-navn → kanonisk app-navn (participants.ts), per app («Håkon M» er
# drammen-Håkon i drammen-ark, men en egen deltaker i alles). Utvid ved behov.
NAME_ALIASES: dict[str, dict[str, str]] = {
    'alles': {
        'Ole Kristian': 'Ole',
        'Kay-Robin': 'Kay Robin',
    },
    'drammen': {
        'Håkon M': 'Håkon',
    },
}


def norm_team(s: str) -> str:
    """Normaliser lagnavn for robust matching: casefold + kun bokstaver/sifre."""
    s = unicodedata.normalize('NFC', s)
    return ''.join(ch for ch in s.casefold() if ch.isalnum())


# --- Spillernavn (--players): normaliser til kanonisk ETTERNAVN slik siden viser dem ---

# Etternavns-partikler som beholdes («Kevin De Bruyne» → «De Bruyne»).
PLAYER_PARTICLES = {'de', 'van', 'der', 'den', 'di', 'da', 'la', 'le', 'el', 'dos', 'del'}

# Diakritisk-strippet nøkkel → kanonisk etternavn slik API-et staver det (matcher auto-fasiten).
# Fanger norsk stavemåte («Håland» → Haaland). Utvid ved behov når nye varianter dukker opp.
PLAYER_ALIASES = {
    'haland': 'Haaland',
    'odegard': 'Ødegaard',
    'sorlot': 'Sørloth',
    'sorloth': 'Sørloth',
    'nusa': 'Nusa',
    'mbappe': 'Mbappé',
}


def strip_dia(s: str) -> str:
    # NFD fjerner diakritikk (é, å, ü …), men IKKE bokstaver med strek (ø/Ø) – map dem manuelt.
    s = s.replace('ø', 'o').replace('Ø', 'O').replace('æ', 'ae').replace('Æ', 'AE')
    return ''.join(ch for ch in unicodedata.normalize('NFD', s) if not unicodedata.combining(ch))


def normalize_player(raw: str) -> str:
    """«Erling Braut Håland» → «Haaland», «kevin de bruyne» → «De Bruyne», «Saka» → «Saka»."""
    parts = raw.strip().split()
    if not parts:
        return ''
    if len(parts) >= 2 and parts[-2].lower() in PLAYER_PARTICLES:
        surname = f'{parts[-2].capitalize()} {parts[-1].capitalize()}'
        key = strip_dia(f'{parts[-2]} {parts[-1]}').casefold()
    else:
        surname = parts[-1].capitalize()
        key = strip_dia(parts[-1]).casefold()
    return PLAYER_ALIASES.get(key, surname)


def parse_score(cell: object) -> tuple[int, int] | None:
    """Tolk et «h-b»-tips. Godtar 1-1, 1–1, 1:1, «1 - 1». Tomt/uleselig → None."""
    if cell is None:
        return None
    m = re.match(r'^\s*(\d+)\s*[-–:]\s*(\d+)\s*$', str(cell))
    return (int(m.group(1)), int(m.group(2))) if m else None


def canonical_names(app: str) -> list[str]:
    src = (REPO_ROOT / 'apps' / app / 'src' / 'data' / 'participants.ts').read_text(encoding='utf-8')
    return re.findall(r"^\s*name:\s*['\"]([^'\"]+)['\"]", src, re.M)


def header_names(cells: list[object], first_col: int) -> dict[int, str]:
    """Kolonneindeks → deltakernavn fra en header-rad (tomme kolonner hoppes over)."""
    out: dict[int, str] = {}
    for i in range(first_col, len(cells)):
        v = cells[i]
        if v is not None and str(v).strip():
            out[i] = str(v).strip()
    return out


def resolve_names(raw: dict[int, str], canon: list[str], app: str) -> dict[int, str]:
    """Map regneark-navn → app-navn via NAME_ALIASES; stopp hardt på ukjente."""
    resolved: dict[int, str] = {}
    unknown: list[str] = []
    aliases = NAME_ALIASES.get(app, {})
    for col, name in raw.items():
        mapped = aliases.get(name, name)
        if mapped in canon:
            resolved[col] = mapped
        else:
            unknown.append(name)
    if unknown:
        sys.exit(
            f'FEIL: ukjente deltakernavn i arket for «{app}»: {unknown}\n'
            f'Gyldige navn: {canon}\n'
            f'Legg inn mapping i NAME_ALIASES i {Path(__file__).name} om navnet er en variant.'
        )
    return resolved


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('xlsx', help='regneark med runde-tips (Resultater + Krydder-ark)')
    ap.add_argument('--app', required=True, choices=['drammen', 'alles'])
    ap.add_argument('--round', required=True, choices=sorted(ROUND_STAGE), help='runde (velger apiId-oppsett; SF/BRONSE/FINALE hentes live fra API-et)')
    ap.add_argument('--k', nargs='*', default=[], help='question-id per Krydder-rad i rekkefølge, f.eks. k4 k5 k6')
    ap.add_argument('--players', nargs='*', default=[], help='id-er blant --k der svaret er spillernavn (normaliseres til etternavn, f.eks. k4)')
    ap.add_argument('-o', '--out', help='output-fil (default: stdout)')
    args = ap.parse_args()

    fixtures = FIXTURES.get(args.round) or fetch_fixtures(args.round)
    by_pair = {frozenset(map(norm_team, pair)): (api_id, pair) for api_id, pair in fixtures.items()}
    canon = canonical_names(args.app)

    wb = load_workbook(args.xlsx, read_only=True, data_only=True)

    # --- Ark «Resultater»: kamptips ---
    ws = wb['Resultater']
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    names = resolve_names(header_names(rows[0], 3), canon, args.app)

    knockout: dict[str, list[dict]] = {}
    matched_ids: set[int] = set()
    flips: list[str] = []
    for row in rows[1:]:
        if not row or row[0] is None or not str(row[0]).strip():
            continue
        home, away = str(row[0]).strip(), str(row[2]).strip()
        key = frozenset((norm_team(home), norm_team(away)))
        if key not in by_pair:
            sys.exit(f'FEIL: kampen «{home} - {away}» finnes ikke i {args.round}-oppsettet: {list(fixtures.values())}')
        api_id, (api_home, _) = by_pair[key]
        matched_ids.add(api_id)
        flipped = norm_team(home) != norm_team(api_home)
        if flipped:
            flips.append(f'{home} - {away} (API: {api_home} hjemme) → tips snus')
        for col, name in names.items():
            score = parse_score(row[col] if col < len(row) else None)
            if score is None:
                continue
            h, a = (score[1], score[0]) if flipped else score
            knockout.setdefault(name, []).append({'apiId': api_id, 'homeGoals': h, 'awayGoals': a})

    # --- Ark «Krydder»: krydder-svar (valgfritt) ---
    bonus: dict[str, dict[str, str]] = {}
    renamed: list[str] = []
    if args.k:
        ws = wb['Krydder']
        krows = [list(r) for r in ws.iter_rows(values_only=True)]
        knames = resolve_names(header_names(krows[0], 2), canon, args.app)
        qrows = [r for r in krows[1:] if r and r[0] is not None and str(r[0]).strip()]
        if len(qrows) != len(args.k):
            sys.exit(f'FEIL: {len(qrows)} krydder-rader i arket, men {len(args.k)} id-er gitt med --k.')
        for qid, row in zip(args.k, qrows):
            for col, name in knames.items():
                v = row[col] if col < len(row) else None
                if v is None or not str(v).strip():
                    continue
                # Heltall fra Excel (3 / 3.0) → «3»; ellers rå tekst (typoer bevares bevisst).
                if isinstance(v, (int, float)) and float(v).is_integer():
                    v = int(v)
                v = str(v).strip()
                # Spillernavn-spørsmål: hvert komma-element → kanonisk etternavn (som på siden).
                if qid in args.players:
                    nv = ', '.join(filter(None, (normalize_player(x) for x in v.split(','))))
                    if nv != v:
                        renamed.append(f'{qid} {name}: «{v}» → «{nv}»')
                    v = nv
                bonus.setdefault(name, {})[qid] = v

    # --- Oppsummering + output ---
    out: dict[str, object] = {}
    if knockout:
        out['knockoutTips'] = knockout
    if bonus:
        out['bonusTips'] = bonus

    n_tips = sum(len(v) for v in knockout.values())
    n_bonus = sum(len(v) for v in bonus.values())
    print(f'-- {args.app} / {args.round}: {len(knockout)} deltakere, {n_tips} kamptips, {n_bonus} krydder-svar', file=sys.stderr)
    missing_matches = set(fixtures) - matched_ids
    if missing_matches:
        print(f'   OBS: kamper uten rad i arket: {sorted(missing_matches)}', file=sys.stderr)
    for f in flips:
        print(f'   snudd: {f}', file=sys.stderr)
    for name in sorted(set(names.values()) | set(n for n in bonus)):
        nt = len(knockout.get(name, []))
        if nt < len(matched_ids):
            print(f'   OBS: {name} mangler {len(matched_ids) - nt} kamptips', file=sys.stderr)
    # Unike svar per krydderspørsmål – kjapp sjekk for rusk («2 kamper» o.l.) før innliming.
    for qid in args.k:
        vals = sorted({v[qid] for v in bonus.values() if qid in v})
        print(f'   {qid}: {vals}', file=sys.stderr)
    for r in renamed:
        print(f'   normalisert: {r}', file=sys.stderr)

    blob = json.dumps(out, ensure_ascii=False, indent=2)
    if args.out:
        Path(args.out).write_text(blob + '\n', encoding='utf-8')
        print(f'   skrev {args.out}', file=sys.stderr)
    else:
        print(blob)


if __name__ == '__main__':
    main()
